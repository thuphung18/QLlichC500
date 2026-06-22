import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

def inspect_xlsx(file_path):
    print(f"\n--- INSPECTING FILE: {file_path} ---")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"Sheet: {sheet_name}")
            sheet = wb[sheet_name]
            # Read first 10 rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx >= 15:
                    break
                row_vals = [str(c).strip() if c is not None else "" for c in row]
                print(f"Row {r_idx + 1}: {row_vals}")
        wb.close()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

inspect_xlsx("e:/Cong_Viec_Chuyen_mon/Flutter_project/QlLich/danh_sach_tai_khoan.xlsx")
inspect_xlsx("e:/Cong_Viec_Chuyen_mon/Flutter_project/QlLich/NV7.xlsx")
