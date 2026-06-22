import sys
import os
import openpyxl
import pyodbc
sys.stdout.reconfigure(encoding='utf-8')

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import get_connection

def update_excel_file():
    xlsx_path = "e:/Cong_Viec_Chuyen_mon/Flutter_project/QlLich/danh_sach_tai_khoan.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found.")
        return
        
    try:
        # 1. Fetch all users from DB
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT u.id, u.username, u.password_hash, u.full_name, u.role, ISNULL(d.name, '')
            FROM dbo.users u
            LEFT JOIN dbo.departments d ON u.department_id = d.id
            ORDER BY d.name, u.full_name
        """)
        db_users = cursor.fetchall()
        cursor.close()
        conn.close()
        
        print(f"Fetched {len(db_users)} users from database to sync to Excel.")
        
        # 2. Load the workbook and worksheet
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb["Danh sách tài khoản"]
        
        # Clear existing data rows (from row 4 onwards)
        max_row = sheet.max_row
        if max_row >= 4:
            sheet.delete_rows(4, max_row - 3)
            
        # 3. Write database users
        for row_idx, user in enumerate(db_users, start=4):
            user_id = str(user[0])
            username = str(user[1])
            password = str(user[2])
            full_name = str(user[3])
            role = str(user[4])
            dept_name = str(user[5])
            
            sheet.cell(row=row_idx, column=1, value=user_id)
            sheet.cell(row=row_idx, column=2, value=username)
            sheet.cell(row=row_idx, column=3, value=password)
            sheet.cell(row=row_idx, column=4, value=full_name)
            sheet.cell(row=row_idx, column=5, value=role)
            sheet.cell(row=row_idx, column=6, value=dept_name)
            
        wb.save(xlsx_path)
        wb.close()
        print(f"Successfully updated {xlsx_path} with {len(db_users)} users.")
        
    except Exception as e:
        print("Error updating Excel file:", e)

if __name__ == "__main__":
    update_excel_file()
